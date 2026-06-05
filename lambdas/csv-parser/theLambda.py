import boto3
import io
import logging
import json
import email
import os
import zipfile
import xml.etree.ElementTree as ET
import csv
import openpyxl
from openpyxl.utils import column_index_from_string
from datetime import datetime

# AWS clients
s3 = boto3.client("s3")
sns_client = boto3.client("sns")

# Logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Config
CSV_BUCKET = os.environ.get("CSV_BUCKET")
NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


# =========================
# New Logic
# =========================
#Idea, we take the first two files and make copies of them and place them inside the s3 bucket. 
#Then for NEW IT contractor report, we take the open and closed sheets and make copies of them and place them inside the s3 bucket.
def process_fileNEW(filename, file_content):
    try:
        if not filename.lower().endswith(".xlsx"):
            logger.info("Skipping unsupported file in process_fileNEW: %s", filename)
            return []

        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
        base_name = filename.rsplit(".", 1)[0]
        clean_base = base_name.split(" 2")[0].split("-2")[0].rstrip("- ")

        results = []

        #For NEW-IT Contractor report, split the two sheets into two seperate excel files Open and Closed
        if filename.startswith("NEW-IT Contractor-VG-Vendor Req Report"):
            for sheet in ("Open", "Closed"):
                logger.info("Processing sheet '%s' (NEW-IT) from: %s", sheet, filename)

                # Extract the sheet into a standalone XLSX
                sheet_bytes = extract_sheet_to_xlsx_bytes(file_content, sheet)
                if not sheet_bytes:
                    logger.warning("No rows or sheet '%s' not found in %s", sheet, filename)
                    continue

                xlsx_key = f"{clean_base}-{sheet}-{timestamp}.xlsx"
                s3.put_object(
                    Bucket=CSV_BUCKET,
                    Key=xlsx_key,
                    Body=sheet_bytes,
                    ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                logger.info("Uploaded XLSX (NEW-IT): %s/%s", CSV_BUCKET, xlsx_key)
                results.append(xlsx_key)

            return results

        # Default behavior: create a timestamped copy of the incoming file in the CSV bucket
        copy_key = f"{clean_base}-copy-{timestamp}.xlsx"
        s3.put_object(
            Bucket=CSV_BUCKET,
            Key=copy_key,
            Body=file_content,
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        logger.info("Uploaded copy: %s/%s", CSV_BUCKET, copy_key)
        results.append(copy_key)
        return results

    except Exception as e:
        logger.error("Error in process_fileNEW %s: %s", filename, e)
        return []
    
def extract_sheet_to_xlsx_bytes(xlsx_bytes, sheet_name):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
        if sheet_name not in wb.sheetnames:
            return None

        src = wb[sheet_name]
        new_wb = openpyxl.Workbook()
        dst = new_wb.active
        dst.title = src.title

        # Copy cells with values and basic styles/number formats
        for row in src.iter_rows():
            for cell in row:
                # Determine column index safely
                col_idx = getattr(cell, 'col_idx', None)
                if col_idx is None:
                    try:
                        col_idx = column_index_from_string(cell.column)
                    except Exception:
                        continue

                dcell = dst.cell(row=cell.row, column=col_idx, value=cell.value)
                try:
                    dcell.number_format = cell.number_format
                    if getattr(cell, 'has_style', False):
                        dcell.font = cell.font
                        dcell.fill = cell.fill
                        dcell.border = cell.border
                        dcell.alignment = cell.alignment
                        dcell.protection = cell.protection
                except Exception:
                    pass

        # Copy column widths
        try:
            for col_letter, dim in src.column_dimensions.items():
                if dim.width is not None:
                    new_wb.active.column_dimensions[col_letter].width = dim.width
        except Exception:
            pass

        # Copy row heights
        try:
            for idx, dim in src.row_dimensions.items():
                if getattr(dim, 'height', None) is not None:
                    new_wb.active.row_dimensions[idx].height = dim.height
        except Exception:
            pass

        # Copy merged cells
        try:
            for merged in src.merged_cells.ranges:
                dst.merge_cells(str(merged))
        except Exception:
            pass

        # Freeze panes
        try:
            if src.freeze_panes:
                dst.freeze_panes = src.freeze_panes
        except Exception:
            pass

        buf = io.BytesIO()
        new_wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    except Exception as e:
        logger.error("Error extracting sheet %s: %s", sheet_name, e)
        return None




# =========================
# XLSX Parsing Helper
# =========================
def col_letter_to_index(col):
    index = 0
    for char in col:
        index = index * 26 + (ord(char.upper()) - ord('A') + 1)
    return index - 1


# =========================
# CSV Parsing
# =========================
def parse_csv_to_rows(csv_bytes):
    try:
        text = csv_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = csv_bytes.decode("latin1")

    reader = csv.reader(io.StringIO(text))
    return list(reader)


# =========================
# SNS Publisher
# =========================
def publish_to_sns(object_key: str):
    topic_arn = os.environ.get("SNS_TOPIC_ARN")

    message = {
        "status": "processed",
        "bucket": CSV_BUCKET,
        "retAddr": object_key
    }

    try:
        response = sns_client.publish(
            TopicArn=topic_arn,
            Message=json.dumps(message),
            Subject="New XLSX File Processed"
        )
        logger.info("Published SNS message: %s", response["MessageId"])

    except Exception as e:
        logger.error("Error publishing to SNS: %s", e)


# =========================
# File Processor
# =========================
SHEET_CONFIG = {
    "ES&F_GR&S Unfilled Requisition Report": [("Sheet1", None)],
    "GR&S Candidate Flow Weekly Report":      [("Sheet1", None)],
    "NEW-IT Contractor-VG-Vendor Req Report": [("Open", "Open"), ("Closed", "Closed")],
}

    


def wipe_buckets():
    s3_resource = boto3.resource("s3")
    bucket = s3_resource.Bucket("rollcall-s3-csv")
    bucket.objects.all().delete()


# =========================
# Lambda Handler
# =========================
def lambda_handler(event, context):
    # Delete all objects before putting news ones in it
    wipe_buckets()

    logger.info(f"Event received: {json.dumps(event, indent=2)}")
    sender_email =json.loads(json.loads(event["Records"][0]["body"])["Message"])["mail"]["source"]
    logger.info(f"Email source detected: {sender_email}")
    total_files = 0
    processed_files = []

    for record in event.get("Records", []):
        sns_message = record["body"]
        message = json.loads(sns_message)["Message"]
        mail_obj = json.loads(message)

        s3_info = mail_obj["receipt"]["action"]
        bucket_name = s3_info["bucketName"]
        object_key = s3_info["objectKey"]

        logger.info("Processing email: %s/%s", bucket_name, object_key)

        email_obj = s3.get_object(Bucket=bucket_name, Key=object_key)
        msg = email.message_from_binary_file(email_obj["Body"])

        for part in msg.walk():
            if "attachment" not in part.get("Content-Disposition", ""):
                continue
            if part.get_content_maintype() == "multipart":
                continue

            filename = part.get_filename()
            if not filename:
                continue

            file_content = part.get_payload(decode=True)
            total_files += 1

            results = process_fileNEW(filename, file_content)
            processed_files.extend(results)

            del file_content

    logger.info("Processed %d/%d files", len(processed_files), total_files)
    publish_to_sns(sender_email)

    return {
        "status": "success",
        "processed_files": processed_files,
        "total_files": total_files
    }